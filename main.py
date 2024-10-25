# load workbook and active sheet
from openpyxl import load_workbook


book = load_workbook("MachineTest.xlsx")
sheet = book.active

# Creating method for adding user
def addUser():
    name = input("Enter Name: ")
    email = input("Enter Email Id: ")
    mob = input("Enter Mobie Number: ")

    # Finding next empty row in excel
    nextRow = sheet.max_row+1

    # Assigning values to row
    sheet[f"A{nextRow}"] = name 
    sheet[f"B{nextRow}"] = email 
    sheet[f"C{nextRow}"] = mob

    # Save the changes to the workbook
    book.save("MachineTest.xlsx")

# Creating method for Showing user saved details
def showDetails():
    # using for loop iterate to all data
    for cell in sheet:
        for data in cell:
            print(data.value)  # Print Excel Data

# Testing of methods
# addUser()
# showDetails()

# It works fine now we will create main method to access these methods using 1 2 3 inputs

def main():
    # Until this is true it will run
    while True:
        print("Make a choice to proceed: ")
        print("Press '1' To Add User: ")
        print("Press '2' To Display Users: ")
        print("Press '3' To Exit: ")

        choice = input("Press any key to continue: ")

        if choice == '1':
            # Here we'll run our 1st method
            addUser()
        elif choice == '2':
             # Here we'll run our 2nd method
            showDetails()
        # For choice 3 we'll exit
        elif choice == '3':
            print("Exit the program")
            # break loop here
            break
        else:
            print("Invalid Choice") #For invalid choice

# Run our main method
main()